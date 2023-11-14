import sys  
sys.path.append(r'./TDD') 
print(sys.path)

import psutil
import os
import psutil
import os
import time
from qiskit import QuantumCircuit
import openpyxl
from gen_benchmark import gen_cir,get_qasm

try:
    from TDD import Ini_TDD,get_unique_table_num
    from TDD_Q import cir_2_tn,get_real_qubit_num
    from QA import QuantumAutomata
    from experiment import image, image_cont_par, image_add_par, generate_ini_state
except:
    raise(ValueError("try to run in QMC directory"))

def initial(quantum_circuit):
    tn,all_indexes=cir_2_tn(quantum_circuit)
    # all_indexes.reverse()
    Ini_TDD(all_indexes)
    return tn

def run(experiment_name:str,qubit_num:int,image_func_name,initial_state_list=[]):


    def run_experiment(gen_cir_func, experiment_func, *args, **kwargs):

        cir=gen_cir_func(qubit_num)
        tn = initial(cir)
        n = get_real_qubit_num(cir)
        t_start = time.time()
        if not initial_state_list:
            ini_state = [generate_ini_state(n,"0"*n).cont()[0]]
        else:
            ini_state = [generate_ini_state(n,state).cont()[0] for state in initial_state_list]

        qa = QuantumAutomata(n,['g'],{'g':tn},ini_state)
        tdd ,max_node = experiment_func(qa, *args, **kwargs)
        print(tdd.node_number())
        pid = os.getpid()
        process = psutil.Process(pid)
        return (time.time()-t_start,max_node,process.memory_info().rss / (1024 * 1024))

    if image_func_name == "image":
        return run_experiment(gen_cir[experiment_name],image)

    elif image_func_name == "image_cont_par":
        if experiment_name == "ghz":
            return run_experiment(gen_cir[experiment_name],image_cont_par,qubit_num//5,qubit_num//5)
        else:
            return run_experiment(gen_cir[experiment_name],image_cont_par, 4, 4)
    else:
        raise(ValueError(f"bad image function name, now is {image_func_name}"))


def generate_data(experiment_name:str,qubit_nums:int,contraction_k:list):
    title = ["qubit number","image","contraction","add partition","one part"]

    time_excel = openpyxl.Workbook()
    table_excel = openpyxl.Workbook()
    time_excel.active.append(title)
    table_excel.active.append(title)

    for i in range(len(qubit_nums)):
        a,b = run(experiment_name,qubit_nums[i],contraction_k[i])
        print(f"{qubit_nums[i]} qubit number of {experiment_name} have been finished\n")
        a.insert(0,qubit_nums[i])
        b.insert(0,qubit_nums[i])

        time_excel.active.append(a)
        table_excel.active.append(b)
        time_excel.save(f"data/a_update_run_time_of_{experiment_name}.xlsx")
        table_excel.save(f"data/a_update_table_number_of_{experiment_name}.xlsx")
    print("all qubit have done\n")

def generate_data_k_add(experiment_name,qubit_num):
    title = ["k","add partition"]

    time_excel = openpyxl.Workbook()
    table_excel = openpyxl.Workbook()
    time_excel.active.append(title)
    table_excel.active.append(title)

    file_name = f"Benchmarks/{experiment_name}_{qubit_num}.qasm"
    cir=QuantumCircuit.from_qasm_file(file_name)

    tn = initial(cir)
    n = get_real_qubit_num(cir)
    ini_state = generate_ini_state(n)
    qa = QuantumAutomata(n,['g'],{'g':tn},[ini_state.cont()])

    t_start=time.time()
    image(qa)
    run_time = time.time()-t_start
    table_num= get_unique_table_num()

    for k in range(1,qubit_num+1):
        a= [k]
        b= [k]

        tn = initial(cir)
        qa=QuantumAutomata(n,['g'],{'g':tn},[ini_state.cont()])

        t_start=time.time()
        image_add_par(tn,qa,k)
        a.append(time.time()-t_start)
        b.append(get_unique_table_num())

        print(f"{qubit_num} qubit number of {experiment_name} with k={k} have been finished\n")

        time_excel.active.append(a)
        table_excel.active.append(b)
        time_excel.save(f"data/run_time_of_{experiment_name}_about_{qubit_num}_qubit_add_partition.xlsx")
        table_excel.save(f"data/table_number_of_{experiment_name}_about_{qubit_num}_qubit_using_add_partition.xlsx")
    print("all qubit have done\n")
    print(f"image time:{run_time},image_table:{table_num}")

def generate_data_k_con(experiment_name,qubit_num,initial_state_list=[]):
    title = ["k_1"]
    title.extend(range(1,qubit_num+1))

    time_excel = openpyxl.Workbook()
    table_excel = openpyxl.Workbook()
    time_excel.active.append(title)
    table_excel.active.append(title)

    # file_name = f"Benchmarks/{experiment_name}__{qubit_num}.qasm"
    cir=gen_cir[experiment_name](qubit_num)
    for k1 in range(1,qubit_num+1):
        a= [k1]
        b= [k1]
        for k2 in range(1,qubit_num+1):
            tn = initial(cir)
            n = get_real_qubit_num(cir)
            tn = initial(cir)
            t_start=time.time()
            if not initial_state_list:
                ini_state = [generate_ini_state(n,"0"*n).cont()[0]]
            else:
                ini_state = [generate_ini_state(n,state).cont()[0] for state in initial_state_list]
            qa=QuantumAutomata(n,['g'],{'g':tn},ini_state)
            image_cont_par(qa,k1,k2)
            a.append(time.time()-t_start)
            b.append(get_unique_table_num())
            print(f"{qubit_num} qubit number of {experiment_name} with k1={k1}, k2={k2} have been finished\n")

        print(f"{qubit_num} qubit number of {experiment_name} with k1={k1} have been finished\n")
        time_excel.active.append(a)
        table_excel.active.append(b)
        time_excel.save(f"data/run__time_of_{experiment_name}_about_{qubit_num}_using_contraction.xlsx")
        table_excel.save(f"data/table__number___of_{experiment_name}_about_{qubit_num}_using_contraction.xlsx")
    print("all qubit have done\n")